"""Generate a styled Excel workbook from the current gold sample_leaders.parquet.

Usage:
    python scripts/generate_sample_excel.py

Output: data/gold/candidates_<sample_size>_cohort_<rule_version>.xlsx
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
import pyarrow.parquet as pq

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.config.settings import GOLD_DIR  # noqa: E402


def _load_sample() -> tuple[pd.DataFrame, dict]:
    sample_df = pq.read_table(GOLD_DIR / "sample_leaders.parquet").to_pandas()
    with open(GOLD_DIR / "sample_manifest.json", encoding="utf-8") as fh:
        manifest = json.load(fh)
    return sample_df, manifest


def _build_display_df(sample_df: pd.DataFrame) -> pd.DataFrame:
    """Select and rename columns for the human-readable Excel sheet."""
    display_cols = {
        "full_name": "Full Name",
        "gender": "Gender",
        "city_size_bucket": "City Size",
        "commune_name": "Commune",
        "dep_code": "Dept.",
        "reg_code": "Region Code",
        "list_nuance": "Nuance",
        "nuance_group": "Nuance Group",
        "is_incumbent": "Incumbent?",
        "score_tour1_pct_expressed": "T1 Vote % (expressed)",
        "score_tour1_rank": "T1 Rank",
        "score_tour2_pct_expressed": "T2 Vote % (expressed)",
        "score_tour2_rank": "T2 Rank",
        "won_final_round": "Won?",
        "advanced_to_tour2": "Advanced T2?",
    }
    display_df = (
        sample_df[list(display_cols.keys())]
        .rename(columns=display_cols)
        .sort_values(["City Size", "Gender", "Full Name"])
        .reset_index(drop=True)
    )
    # city_size sort order: large â†’ medium â†’ small
    size_order = {"large": 0, "medium": 1, "small": 2}
    display_df["_sort_size"] = display_df["City Size"].map(size_order)
    display_df = (
        display_df.sort_values(["_sort_size", "Gender", "Full Name"])
        .drop(columns=["_sort_size"])
        .reset_index(drop=True)
    )
    return display_df


# â”€â”€ Colour palette â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Row background: city-size bucket Ã— gender  (ARGB hex, no leading #)
_FILL = {
    ("large", "F"): "FFD6E4F0",  # light blue
    ("large", "M"): "FFBBD6E8",  # medium blue
    ("medium", "F"): "FFD6F0D6",  # light green
    ("medium", "M"): "FFB3E0B3",  # medium green
    ("small", "F"): "FFFFF0C8",  # light yellow
    ("small", "M"): "FFFFD980",  # medium yellow
}
_HEADER_FILL = "FF2F5496"  # dark blue
_HEADER_FONT = "FFFFFFFF"  # white


def _write_excel(display_df: pd.DataFrame, manifest: dict, out_path: Path) -> None:
    """Write the styled Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is not installed. Run: python -m pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    by_gender = manifest.get("by_gender", {})
    by_city_size = manifest.get("by_city_size", {})
    gender_breakdown = f"{by_gender.get('F', 0)} / {by_gender.get('M', 0)}"
    city_size_breakdown = " / ".join(
        str(by_city_size.get(bucket, 0)) for bucket in ("large", "medium", "small")
    )

    # â”€â”€ Sheet 1: Cohort â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws = wb.active
    rule_version = manifest.get("sampling_rule_version", "cohort")
    ws.title = f"Cohort_{rule_version}"[:31]

    header_fill = PatternFill(fill_type="solid", fgColor=_HEADER_FILL)
    header_font = Font(bold=True, color=_HEADER_FONT, size=11)
    thin_side = Side(style="thin", color="FFAAAAAA")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    # Write header
    for col_idx, col_name in enumerate(display_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border

    # Write data rows
    for row_idx, row in display_df.iterrows():
        city_size = row["City Size"]
        gender = row["Gender"]
        row_fill = PatternFill(
            fill_type="solid",
            fgColor=_FILL.get((city_size, gender), "FFFFFFFF"),
        )
        for col_idx, value in enumerate(row, start=1):
            # pandas NA/NaT cannot be written to openpyxl â€” coerce to None
            if pd.isna(value) if not isinstance(value, (str, bool)) else False:
                value = None
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # Right-align numeric columns
            col_name = display_df.columns[col_idx - 1]
            if col_name in ("T1 Vote % (expressed)", "T2 Vote % (expressed)"):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, float):
                    cell.number_format = "0.00%"
                    cell.value = value / 100 if value is not None else None
            elif col_name in ("T1 Rank", "T2 Rank"):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    col_widths = {
        "Full Name": 28,
        "Gender": 8,
        "City Size": 10,
        "Commune": 28,
        "Dept.": 7,
        "Region Code": 12,
        "Nuance": 10,
        "Nuance Group": 14,
        "Incumbent?": 11,
        "T1 Vote % (expressed)": 16,
        "T1 Rank": 8,
        "T2 Vote % (expressed)": 16,
        "T2 Rank": 8,
        "Won?": 7,
        "Advanced T2?": 12,
    }
    for col_idx, col_name in enumerate(display_df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(
            col_name, 14
        )

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # â”€â”€ Sheet 2: Legend â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws_legend = wb.create_sheet("Legend")
    ws_legend.column_dimensions["A"].width = 22
    ws_legend.column_dimensions["B"].width = 40

    legend_title_font = Font(bold=True, size=12)
    legend_rows = [
        ("Colour legend", "", None),
        ("Large city â€“ Female", "", _FILL[("large", "F")]),
        ("Large city â€“ Male", "", _FILL[("large", "M")]),
        ("Medium city â€“ Female", "", _FILL[("medium", "F")]),
        ("Medium city â€“ Male", "", _FILL[("medium", "M")]),
        ("Small city â€“ Female", "", _FILL[("small", "F")]),
        ("Small city â€“ Male", "", _FILL[("small", "M")]),
        ("", "", None),
        ("Sampling rule", manifest.get("sampling_rule_version", ""), None),
        ("Total candidates", str(manifest.get("total_sampled", len(display_df))), None),
        ("Female / Male", gender_breakdown, None),
        ("Large / Medium / Small", city_size_breakdown, None),
        ("Distinct regions", str(manifest.get("distinct_regions", "")), None),
        (
            "Max per region",
            str(manifest["hard_constraints"].get("max_candidates_per_region", "")),
            None,
        ),
        ("Random seed", str(manifest.get("random_seed", 42)), None),
        ("Run ID", manifest.get("run_id", ""), None),
        ("Created at", manifest.get("created_at", ""), None),
    ]

    for r_idx, (label, value, fill_hex) in enumerate(legend_rows, start=1):
        cell_a = ws_legend.cell(row=r_idx, column=1, value=label)
        cell_b = ws_legend.cell(row=r_idx, column=2, value=value)
        if r_idx == 1:
            cell_a.font = legend_title_font
        if fill_hex:
            row_fill = PatternFill(fill_type="solid", fgColor=fill_hex)
            cell_a.fill = row_fill
            cell_b.fill = row_fill

    wb.save(out_path)
    print(f"Saved: {out_path}")


def main() -> None:
    sample_df, manifest = _load_sample()
    display_df = _build_display_df(sample_df)
    rule_version = manifest.get("sampling_rule_version", "cohort")
    total_sampled = int(manifest.get("total_sampled", len(display_df)))
    out_path = GOLD_DIR / f"candidates_{total_sampled}_cohort_{rule_version}.xlsx"
    _write_excel(display_df, manifest, out_path)

    # Quick summary
    sep = "-" * 55
    print(f"\n{sep}")
    print(f"  Rule version : {manifest['sampling_rule_version']}")
    print(f"  Total        : {len(display_df)} candidates")
    print(f"  Gender       : {manifest['by_gender']}")
    print(f"  City size    : {manifest['by_city_size']}")
    print(f"  Regions      : {manifest['distinct_regions']} distinct")
    print(sep)


if __name__ == "__main__":
    main()
